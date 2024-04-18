#계산된 결과
from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active


for row in ws.values:
  for cell in row:
    print(cell)


#수식그래도 가져옴
#from openpyxl import load_workbook
#wb = load_workbook("sample_formula.xlsx")
#ws = wb.active

#for row in ws.values:
#  for cell in row:
#    print(cell)
