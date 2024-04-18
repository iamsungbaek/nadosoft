from openpyxl import load_workbook
wb = load_workbook("Sample2.xlsx")
ws = wb.active


# 번호 영어 수학 --엑셀파일에 있는 글자
# 번호 국어 영어 수학 -- 국어 추가 및 이동
#ws.move_range("B1:C11", rows=0, cols=1)

#ws["B1"].value = "국어" # B1셀에 국어 입력

#wb.save("sample_국어.xlsx")

ws.move_range("c1:c11", rows=5, cols=-1)

wb.save("sample_국어.xlsx")