from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로은 시트 기본 이름으로 생성
ws.title = "MySheet1"
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("Your1Sheet")
ws2 = wb.create_sheet("NewSheet", 2) # 2번쨰 index에 시트 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근

print(wb.sheetnames) # 모든 시트명 확인

#시트복사
new_ws["A1"] = "TEST"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("Sample3.xlsx")
#wb.close()