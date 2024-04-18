from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("Sample2.xlsx")
ws = wb.active

# 엑섹의 셀
a1 = ws["A1"] #번호
b1 = ws["B1"] #영어
c1 = ws["C1"] #수학

ws.column_dimensions["A"].width = 5  # A열의 너비를 5로 설정
ws.row_dimensions[1].height = 50 # 1행의 너비를 50로 설정


#스타일적용
a1.font = Font(color ="FF0000", italic = True, bold = True)
b1.font = Font(color ="CC33FF", name="Arial", strike = True)
c1.font = Font(color ="FFCC66", size = 20, underline= "single")

#테두리적용
thin_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

#배경색(90점 넘는 셀에 대해서 초록색 적용)
for row in ws.rows: #모든 셀 적용
  for cell in row:
      cell.alignment = Alignment(horizontal="center", vertical="center") #각 셀에 대해서 정렬

      if cell.column == 1: # A 번호열은 제외
          continue

       # cell 이 정수형 데이터이고 90점보다 높으면
      if isinstance(cell.value, int) and cell.value > 90:
           cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
           cell.font = Font(color="FF0000")

# 틀 고정
ws.freeze_panes = "B2"

wb.save("sample_style.xlsx")

