# 셀병합한거 풀기
from openpyxl import load_workbook
wb = load_workbook("sample_셀병합.xlsx")
ws = wb.active

#셀 병합하기
ws.unmerge_cells("B2:D2")

wb.save("sample_un셀병합.xlsx")



#셀 병합하기
#from openpyxl import Workbook
#wb = Workbook()
#ws = wb.active

#ws.merge_cells("B2:D2")
#ws["B2"].value = "병합하기"

#wb.save("sample_셀병합.xlsx")


