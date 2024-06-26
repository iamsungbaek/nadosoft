from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호","영어","수학"]) # A B C ....
for i in range(1,11):
    ws.append([i, randint(0,100), randint(0,100)])

col_B = ws["B"] # 영어 컬럼을 가져오기
#print(col_B)
for cell in col_B:
    print(cell.value)

col_range =  ws["B:C"] #영어 수학 컬럼 함께 가져오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1] # 1번째 row 만 가져오기
# for cell in row_title:
#    print(cell.value)

#row_range = ws[2:6] # 2번재 줄에서 6번째 줄까지 가지고 오기(1번째 제외)
#for rows in row_range:
#    for cell in rows:
#        print(cell.value, end = " ")
#    print()

from openpyxl.utils.cell import coordinate_from_string



#row_range = ws[2:ws.max_row] # 2번재 줄부터 끝까지
#for rows in row_range:
#    for cell in rows:
#        #print(cell.value, end = " ")
#        #print(cell.coordinate, end = " ")
#        xy = coordinate_from_string(cell.coordinate)
#       # print(xy, end = " ")
#        print(xy[0], end = "") # A
#        print(xy[1], end = " ") # 1
#    print() # 줄바꿈을 위해서

##print(tuple(ws.rows)) #전체 rows

for row in tuple(ws.rows):
    print(row[1].value)

##print(tuple(ws.columns)) #전체 columns

for column in tuple(ws.columns):
    print(column[0].value)

for row in ws.iter_rows(): #전체 rows
    print(row[1].value)

for column in ws.iter_cols():
    print(column[2].value)


# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row = 2, max_row =11, min_col = 2, max_col = 3): # 열
    #print(row[0].value,row[1].value)
    print(row)

for col in ws.iter_cols(min_row = 2, max_row =11, min_col = 2, max_col = 3): # 행
    print(col)

wb.save("Sample2.xlsx")
