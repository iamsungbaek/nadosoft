from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("sample1.xlsx") # sample1.xlsx 파일에서 wb을 불러옴
ws = wb.active

#cell 데이터 불러오기
# for x in range(1,11):
#   for y in range(1,11):
#     print(ws.cell(row=x, column=y).value, end=" ")
#   print()

#cell 갯수를 모를때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()