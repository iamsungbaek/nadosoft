import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today() # 오늘날짜
ws["A2"] = "=sum(1,2,3)"             # 1+2+3 = 6
ws["A3"] = "=AVERAGE(1,2,3)"         # 평균

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "SUM(A4:A5)"


wb.save("sample_formula.xlsx")
